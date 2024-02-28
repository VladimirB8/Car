from PIL import Image
from imageai.Detection import ObjectDetection
import os
import tempfile
from openpyxl import Workbook

def crop_image(image_path):
    image = Image.open(image_path)
    width, height = image.size
    left = width // 2
    top = 0
    right = width
    bottom = height
    cropped_image = image.crop((left, top, right, bottom))
    return cropped_image

def detect_objects_in_images(image_folder):
    execution_path = os.getcwd()
    total_images = len([f for f in os.listdir(image_folder) if f.endswith('.jpg') or f.endswith('.png')])
    processed_images = 0
    total_cars_all_images = 0

    # загружаем модель YOLOv3
    detector = ObjectDetection()
    detector.setModelTypeAsYOLOv3()
    detector.setModelPath(os.path.join(execution_path, "yolov3.pt"))
    detector.loadModel()

    custom = detector.CustomObjects(car=True)

    wb = Workbook()
    ws = wb.active
    ws.append(["Имя изображения", "Количество машин"])

    for image_file in os.listdir(image_folder):
        if image_file.endswith(".jpg") or image_file.endswith(".png"):
            processed_images += 1
            print(f"Обработано изображение {processed_images} из {total_images}")

            image_path = os.path.join(image_folder, image_file)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_image:
                temp_image_path = temp_image.name

                cropped_image = crop_image(image_path)
                cropped_image.save(temp_image_path)

                # обнаруживаем объекты на обрезанном изображении
                detections = detector.detectObjectsFromImage(custom_objects=custom, input_image=temp_image_path, minimum_percentage_probability=5)

                # считаем количество машин на изображении
                total_cars = sum(1 for obj in detections if obj["name"] == "car")
                total_cars_all_images += total_cars

                ws.append([image_file, total_cars])

            os.unlink(temp_image_path)

    ws.cell(row=1, column=4, value="Общее количество машин на всех изображениях")
    ws.cell(row=2, column=4, value=total_cars_all_images)
    wb.save("результаты.xlsx")

    print("Обработка завершена.")

image_folder = input("Введите путь к директории с изображениями: ")

detect_objects_in_images(image_folder)
